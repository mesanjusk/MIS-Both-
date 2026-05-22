#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()

RED    = "C0392B"
ORANGE = "E67E22"
GREEN  = "27AE60"
BLUE   = "2980B9"
PURPLE = "8E44AD"
DARK   = "1A1A2E"
WHITE  = "FFFFFF"
LGREY  = "F2F3F4"
MGREY  = "BDC3C7"

def hfill(hex_):   return PatternFill("solid", fgColor=hex_)
def bold(sz=11, colour=WHITE): return Font(bold=True, size=sz, color=colour)
def normal(sz=10, colour="333333"): return Font(size=sz, color=colour)
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)
def thin_border():
    s = Side(style="thin", color=MGREY)
    return Border(left=s, right=s, top=s, bottom=s)

def header_row(ws, row, cols_text, fill_hex, font_col=WHITE, height=28):
    ws.row_dimensions[row].height = height
    for col, text in enumerate(cols_text, 1):
        c = ws.cell(row=row, column=col, value=text)
        c.fill  = hfill(fill_hex)
        c.font  = bold(10, font_col)
        c.alignment = center()
        c.border = thin_border()

def section_header(ws, row, text, fill_hex, span, height=22):
    ws.row_dimensions[row].height = height
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.fill  = hfill(fill_hex)
    c.font  = bold(11, WHITE)
    c.alignment = left()
    c.border = thin_border()

def data_row(ws, row, values, alt=False):
    ws.row_dimensions[row].height = 20
    bg = LGREY if alt else WHITE
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill   = hfill(bg)
        c.font   = normal()
        c.border = thin_border()
        c.alignment = left() if col > 1 else center()

def add_dropdown(ws, col_letter, first_row, last_row, formula):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    dv.sqref = f"{col_letter}{first_row}:{col_letter}{last_row}"
    ws.add_data_validation(dv)


# ══════════════════════════════════════════════════════════════════
# SHEET 1 – PAGES / COMPONENTS TO DELETE
# ══════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "1. Delete Dead Code"
ws1.freeze_panes = "A3"
ws1.sheet_view.showGridLines = False

for letter, width in [("A",5),("B",42),("C",22),("D",40),("E",15),("F",15)]:
    ws1.column_dimensions[letter].width = width

header_row(ws1, 1,
    ["#", "File / Feature", "Area", "Why Remove", "Approve", "Priority"],
    DARK, height=30)

items_delete = [
    ("FE Pages",),
    (1,  "Pages/home.jsx  *** UNROUTED — dead code ***",
         "Home/Landing",
         "/home route serves Dashboard.jsx — home.jsx is not connected to ANY route. 181 lines with 2-sec artificial setTimeout, mixes attendance+tasks+orders. DELETE; Dashboard.jsx becomes the single home screen.",
         "", "High"),
    (2,  "Components/RightSidebar.jsx",
         "Navigation",
         "Duplicates left nav; confusing rail buttons — move 5 actions to top navbar instead.",
         "", "High"),
    (3,  "Pages/FlowBuilderPage.jsx",
         "Automation",
         "Visual flow builder too complex; team of 4 has no time to design automations.",
         "", "High"),
    (4,  "Pages/BankReconciliation.jsx",
         "Accounts",
         "Enterprise feature; wrong scale for 10 orders/day.",
         "", "High"),
    (5,  "Pages/TrialBalance.jsx",
         "Accounts",
         "Accountant-level tool; not needed daily.",
         "", "High"),
    (6,  "Pages/OpeningBalance.jsx",
         "Accounts",
         "One-time setup page; not needed after initial entry.",
         "", "Medium"),
    (7,  "Pages/GmailAccounts.jsx",
         "Email",
         "WhatsApp covers all customer communication; Gmail adds complexity.",
         "", "High"),
    (8,  "Pages/EmailCompose.jsx",
         "Email",
         "Same — consolidated into WhatsApp.",
         "", "High"),
    (9,  "Pages/EmailHistory.jsx",
         "Email",
         "Same — consolidated into WhatsApp.",
         "", "High"),
    (10, "Pages/CallLogs.jsx",
         "CRM",
         "Not core to printing business ops; rarely used.",
         "", "Medium"),
    (11, "Pages/DiaryUpload.jsx",
         "Misc",
         "Non-essential daily upload feature.",
         "", "Medium"),
    (12, "Pages/SopPage.jsx",
         "Ops",
         "Replace with simple daily checklist embedded in dashboard.",
         "", "Medium"),
    (13, "Pages/WorkflowTemplates.jsx",
         "Automation",
         "Replaced by fixed pipeline; team doesn't need custom workflow editing.",
         "", "High"),
    (14, "Pages/UpiPayment.jsx  (standalone)",
         "Accounts",
         "Merge UPI collection into dashboard quick-action widget instead.",
         "", "Medium"),
    (15, "Pages/addTransaction1.jsx",
         "Accounts",
         "Replace with inline quick-add in simple accounts log.",
         "", "Medium"),
    (16, "Pages/AttendanceReport.jsx  (duplicate)",
         "HR",
         "Keep AllAttandance.jsx; remove duplicate report.",
         "", "Low"),
    (17, "Pages/purchaseOrder.jsx",
         "Procurement",
         "Overkill at current scale; add note on vendor payment instead.",
         "", "Medium"),
    ("Legacy WhatsApp (Baileys)",),
    (18, "Pages/WhatsApp*.jsx  (all Baileys pages)",
         "WhatsApp",
         "Unofficial QR-based; fragile; Cloud API replaces all of these.",
         "", "High"),
    (19, "Components/whatsapp/ (legacy folder)",
         "WhatsApp",
         "All legacy WA components; superseded by whatsappCloud/ folder.",
         "", "High"),
    ("Reports — consolidate 17 into 3",),
    (20, "Reports/allTransaction.jsx",
         "Reports",
         "5 transaction views collapse into 1 smart Collections report.",
         "", "High"),
    (21, "Reports/allTransaction4D.jsx",
         "Reports",
         "Same — fold into Collections report.",
         "", "High"),
    (22, "Reports/allBills.jsx  (879-line version)",
         "Reports",
         "Rebuild as simplified Bills tab inside Collections report.",
         "", "High"),
    (23, "Reports/agingReport.jsx",
         "Reports",
         "Overkill; receivables aging not actionable daily for small team.",
         "", "Medium"),
    (24, "Reports/priorityReport.jsx",
         "Reports",
         "Admin-level; low daily use.",
         "", "Low"),
    (25, "Reports/taskReport.jsx",
         "Reports",
         "Tasks visible on dashboard; separate report not needed.",
         "", "Low"),
    (26, "Reports/userReport.jsx",
         "Reports",
         "Admin-level; low daily use.",
         "", "Low"),
    (27, "Reports/allDelivery.jsx",
         "Reports",
         "Fold delivery tracking into Orders report.",
         "", "Medium"),
    (28, "Reports/billUpdate.jsx",
         "Reports",
         "Bill status updates happen from Kanban; standalone page redundant.",
         "", "Medium"),
    (29, "Reports/paymentReport.jsx",
         "Reports",
         "Merge into Collections report.",
         "", "Medium"),
    ("Backend routes + models to remove",),
    (30, "MISBackend/routes/Baileys.js",
         "Backend",
         "Remove with Baileys frontend.",
         "", "High"),
    (31, "MISBackend/services/baileysService.js",
         "Backend",
         "Same.",
         "", "High"),
    (32, "MISBackend/repositories/baileysAuthState.js",
         "Backend",
         "Same.",
         "", "High"),
    (33, "MISBackend/routes/Gmail.js",
         "Backend",
         "Remove with Gmail frontend.",
         "", "High"),
    (34, "MISBackend/routes/BankStatement.js",
         "Backend",
         "Remove with BankReconciliation frontend.",
         "", "High"),
    (35, "MISBackend/routes/workflowTemplate.js",
         "Backend",
         "Remove with WorkflowTemplates frontend.",
         "", "High"),
    (36, "MISBackend/routes/PurchaseOrder.js",
         "Backend",
         "Remove with purchaseOrder frontend.",
         "", "Medium"),
    (37, "MISBackend/repositories/flow.js + flowSession.js",
         "Backend",
         "Remove with FlowBuilder frontend.",
         "", "High"),
]

row = 2
alt = False
for item in items_delete:
    if len(item) == 1:
        section_header(ws1, row, f"  ▶  {item[0]}", BLUE, 6)
        row += 1
        continue
    data_row(ws1, row, list(item), alt)
    c = ws1.cell(row=row, column=5)
    c.fill = hfill("D5F5E3"); c.font = bold(10, "1E8449"); c.alignment = center()
    alt = not alt
    row += 1

add_dropdown(ws1, "E", 2, row - 1, '"Yes,Skip,Later"')


# ══════════════════════════════════════════════════════════════════
# SHEET 2 – COMPONENTS TO REWRITE / SIMPLIFY
# ══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2. Rewrite & Simplify")
ws2.freeze_panes = "A3"
ws2.sheet_view.showGridLines = False

for letter, width in [("A",5),("B",40),("C",16),("D",14),("E",44),("F",14),("G",14)]:
    ws2.column_dimensions[letter].width = width

header_row(ws2, 1,
    ["#","File","Current Lines","Target Lines","Key Changes","Approve","Priority"],
    DARK, height=30)

items_rewrite = [
    ("HOME / LANDING PAGE",),
    (1,  "Pages/home.jsx  →  DELETE (see Sheet 1 row 1)",
         "181", "0",
         "home.jsx is NOT the /home route — Dashboard.jsx is. home.jsx is dead code. "
         "After cleanup, ROUTES.HOME ('/home') continues to serve the new simplified Dashboard.jsx. "
         "No separate 'home' page needed.",
         "", "High"),

    ("NAVIGATION",),
    (2,  "Components/Sidebar.jsx",
         "288", "~80",
         "6 items flat list: Dashboard / Orders / Customers / Accounts / Reports / Settings (admin only). "
         "Remove all groups and collapsible sections. Delete sidebarMenu.jsx config file entirely.",
         "", "High"),
    (3,  "Components/TopNavbar.jsx",
         "100+", "~80",
         "Add [+ New Order] primary button. Move 5 right-sidebar quick actions here. Remove right sidebar dependency.",
         "", "High"),
    (4,  "Components/Footer.jsx  (mobile bottom nav)",
         "50", "~40",
         "4 tabs only: Dashboard / Orders / + New Order / Accounts. Remove Tasks and Chat tabs.",
         "", "Medium"),

    ("DASHBOARD  (= the /home route)",),
    (5,  "Pages/Dashboard.jsx  — serves ROUTES.HOME = '/home'",
         "754", "~200",
         "This IS the home screen after login. "
         "New layout: 5 KPI tiles (Today Orders, Ready, Overdue, Collected ₹, Pending ₹) + "
         "'Needs Attention' list (max 5 stuck/overdue orders with [Move] button) + "
         "Stage pipeline count bar. "
         "REMOVE: resizable panels, customization drawer, attendance section, user tasks section, design files widget.",
         "", "High"),
    (6,  "Components/dashboard/  (folder)",
         "~300", "~80",
         "Keep SummaryCard + QuickActions only. "
         "Delete: CrmSidebarPanel, ActionList, RoleWidget, DesignFilesWidget. "
         "Simplify UpiCollectionSection to a single inline button.",
         "", "High"),

    ("ORDER MANAGEMENT",),
    (7,  "Pages/addOrder1.jsx",
         "1726", "~180",
         "6 fields only: Customer (autocomplete), Items/Qty/Rate table (max 5 rows), Due Date (default +3 days), "
         "Priority (Low/Normal/Urgent chip), Advance Paid (yes/no + amount). "
         "2 save buttons: [Save as Enquiry] and [Save as Order]. "
         "REMOVE: vendor section (auto-assign on backend), task groups, Google Drive, date override, WA toggle (always auto-send).",
         "", "High"),
    (8,  "Pages/OrderUpdate.jsx",
         "710", "~150",
         "Read-only order detail. Editable fields: due date, priority, items qty/rate, add note. "
         "Stage advance moves to Kanban board only.",
         "", "High"),
    (9,  "Pages/OrderKanban.jsx + Components/orders/OrderCard.jsx",
         "~300", "~150",
         "Keep Kanban paradigm. Simpler card: #order · Customer | Item summary | Due: X days | ₹ amount | [→ Next Stage] button. "
         "8 columns = full pipeline. Remove drag-drop; use [→ Stage] button only.",
         "", "High"),

    ("ACCOUNTS  (Simplify)",),
    (10, "Pages/DayBook.jsx",
         "1066", "~120",
         "Replace double-entry complexity with Income/Expense log. "
         "Two tabs: Income (auto-created when order paid) | Expenses (manual). "
         "Table: Date | Type | Amount | Category | Order# | Note. "
         "Totals at top: Income ₹X — Expenses ₹Y = Profit ₹Z.",
         "", "High"),

    ("REPORTS  (Rebuild 17 → 3)",),
    (11, "Reports/allOrder.jsx  →  New: Orders Report",
         "712", "~200",
         "Table + filters: date range, stage, customer. "
         "Columns: Order# | Customer | Items | Amount | Stage | Due Date | Assigned. "
         "Export Excel. Replaces: allOrder, AllOrderTableView, allDelivery.",
         "", "High"),
    (12, "New file: Reports/collectionsReport.jsx",
         "—", "~200",
         "Income + expenses by date. Summary: Total In | Total Out | Net. "
         "Replaces: allBills, allTransaction*, paymentReport. Export Excel.",
         "", "High"),
    (13, "Reports/customerReport.jsx",
         "303", "~150",
         "Minor cleanup: remove unused columns. Keep search, group filter, edit/delete.",
         "", "Low"),

    ("BACKEND",),
    (14, "controllers/orderLifecycleController.js",
         "~200", "~180",
         "Add auto-WhatsApp send after each stage change "
         "using existing whatsappCloudService.sendTemplate(). Reuse whatsappTemplates.js config.",
         "", "High"),
    (15, "services/messageScheduler.js",
         "~150", "~200",
         "Add two cron jobs: (a) daily 10am payment reminder for delivered-unpaid orders; "
         "(b) 7pm owner summary WhatsApp with today's stats.",
         "", "High"),
    (16, "App.jsx  (router)",
         "66 routes", "~35 routes",
         "Remove routes for all deleted pages. Keep ~35 active routes.",
         "", "High"),
]

row = 2
alt = False
for item in items_rewrite:
    if len(item) == 1:
        section_header(ws2, row, f"  ▶  {item[0]}", PURPLE, 7)
        row += 1
        continue
    ws2.row_dimensions[row].height = 22
    data_row(ws2, row, list(item), alt)
    c = ws2.cell(row=row, column=6)
    c.fill = hfill("D5F5E3"); c.font = bold(10,"1E8449"); c.alignment = center()
    alt = not alt
    row += 1

add_dropdown(ws2, "F", 2, row - 1, '"Yes,Skip,Later"')


# ══════════════════════════════════════════════════════════════════
# SHEET 3 – AUTOMATION TO ADD
# ══════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3. Automation to Add")
ws3.freeze_panes = "A3"
ws3.sheet_view.showGridLines = False

for letter, width in [("A",5),("B",30),("C",42),("D",32),("E",18),("F",14),("G",14)]:
    ws3.column_dimensions[letter].width = width

header_row(ws3, 1,
    ["#","Feature","What It Does","Files to Change","Trigger","Approve","Priority"],
    DARK, height=30)

items_auto = [
    ("WHATSAPP AUTO-MESSAGES  (Backend — zero human action)",),
    (1,  "Order confirmation WA",
         "On order creation: auto-send 'Order #N received, amount Rs.X, due date Y' to customer.",
         "orderService.js (create hook)", "Order created", "", "High"),
    (2,  "Stage-change WA message",
         "Auto-send WA to customer when order moves to: approved / design / ready / delivered. "
         "Messages defined in whatsappTemplates.js.",
         "orderLifecycleController.js + whatsappTemplates.js", "PATCH /stage", "", "High"),
    (3,  "Payment reminder WA",
         "Daily 10am: find all delivered-unpaid orders older than 2 days. "
         "Send WA: 'Your order #N of Rs.X is pending payment.'",
         "messageScheduler.js", "Cron 10:00", "", "High"),
    (4,  "Daily owner summary WA",
         "7pm every day: aggregate today's stats (new orders, stage moves, collections, overdue count). "
         "Send WhatsApp summary to owner's number (configured in BusinessOps).",
         "messageScheduler.js", "Cron 19:00", "", "High"),
    (5,  "Enquiry follow-up WA",
         "If enquiry stays in 'quoted' stage > 24h, auto-send follow-up message to customer.",
         "messageScheduler.js", "Cron daily 11:00", "", "Medium"),

    ("ORDER AUTOMATION  (Backend)",),
    (6,  "Auto vendor assignment",
         "On order creation, auto-assign preferred vendor to each production step "
         "based on item type — no manual vendor selection needed.",
         "orderService.js + vendorService.js  (already exists, just not wired)", "Order created", "", "High"),
    (7,  "Auto task creation",
         "Auto-create designer task when order moves to 'design' stage. "
         "(Already partly built in orderLifecycleController — complete the wiring.)",
         "orderLifecycleController.js", "Stage = design", "", "Medium"),
    (8,  "SLA / stuck-order alert",
         "Flag order in dashboard 'Needs Attention' section if it stays in the same stage > 2 days.",
         "dashboardSummaryController.js + /api/dashboard/stuck-orders endpoint", "Dashboard load", "", "Medium"),

    ("ACCOUNTS AUTOMATION",),
    (9,  "Auto income entry",
         "When order is marked 'paid', auto-create income transaction. "
         "No manual journal entry needed.",
         "orderLifecycleController.js + transactionService.js", "Stage = paid", "", "High"),
    (10, "Auto PDF invoice on delivery",
         "When stage moves to 'delivered', auto-generate PDF invoice and attach to order record.",
         "orderLifecycleController.js + InvoiceModal.jsx (existing)", "Stage = delivered", "", "Medium"),
]

row = 2
alt = False
for item in items_auto:
    if len(item) == 1:
        section_header(ws3, row, f"  ▶  {item[0]}", GREEN, 7)
        row += 1
        continue
    ws3.row_dimensions[row].height = 22
    data_row(ws3, row, list(item), alt)
    c = ws3.cell(row=row, column=6)
    c.fill = hfill("D5F5E3"); c.font = bold(10,"1E8449"); c.alignment = center()
    alt = not alt
    row += 1

add_dropdown(ws3, "F", 2, row - 1, '"Yes,Skip,Later"')


# ══════════════════════════════════════════════════════════════════
# SHEET 4 – BACKEND DB / MODEL CLEANUP
# ══════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("4. DB & Backend Cleanup")
ws4.freeze_panes = "A3"
ws4.sheet_view.showGridLines = False

for letter, width in [("A",5),("B",36),("C",18),("D",42),("E",14),("F",14)]:
    ws4.column_dimensions[letter].width = width

header_row(ws4, 1,
    ["#","Collection / Code","Action","Why","Approve","Priority"],
    DARK, height=30)

items_db = [
    ("MONGODB COLLECTIONS — REMOVE",),
    (1,  "Enquiry collection",
         "Delete",
         "All enquiries now stored as Orders with stage=enquiry. Enquiry table is dead code.",
         "", "High"),
    (2,  "Flow + FlowSession",
         "Delete",
         "Removing flow builder frontend; collections unused.",
         "", "High"),
    (3,  "BaileysAuthState + BaileysMessage",
         "Delete",
         "Removing Baileys WhatsApp; collections unused.",
         "", "High"),
    (4,  "EmailHistory + GmailAccount",
         "Delete",
         "Removing Gmail integration.",
         "", "High"),
    (5,  "BankStatement",
         "Delete",
         "Removing bank reconciliation.",
         "", "High"),
    (6,  "PurchaseOrder",
         "Delete",
         "Removing purchase order module.",
         "", "Medium"),
    (7,  "CallLogs",
         "Delete",
         "Removing call logs page.",
         "", "Medium"),
    (8,  "DiaryDraft",
         "Delete",
         "Removing diary upload.",
         "", "Medium"),

    ("ORDER MODEL — FIELD CLEANUP",),
    (9,  "Order.Status[]  (array field)",
         "Remove",
         "Legacy field; replaced by Order.stage (string). Every query must handle both — confusing and error-prone.",
         "", "High"),
    (10, "Order.vendorCustomerUuid",
         "Remove",
         "Old FK; standardise to vendorUuid everywhere.",
         "", "Medium"),
    (11, "Order.workflowSteps[]",
         "Audit",
         "Check if actively used; if fewer than 20% of orders have data, remove entirely.",
         "", "Medium"),

    ("CODE QUALITY",),
    (12, "API naming: /GetOrderList, /GetDeliveredList",
         "Rename to kebab-case",
         "Inconsistent with REST conventions; rename to /orders, /delivered-orders.",
         "", "Low"),
    (13, "Error response format inconsistency",
         "Standardize",
         "Mix of {success,message} and {status,message}; pick one schema and apply everywhere.",
         "", "Low"),
    (14, "Hardcoded 'Sai' as default assignee",
         "Remove",
         "Replace with null or first available user; hardcoded names break on new installs.",
         "", "Medium"),
    (15, "Pagination on /api/orders/all-data",
         "Add",
         "Missing pagination causes slow load; will get worse as order count grows.",
         "", "High"),
]

row = 2
alt = False
for item in items_db:
    if len(item) == 1:
        section_header(ws4, row, f"  ▶  {item[0]}", ORANGE, 6)
        row += 1
        continue
    ws4.row_dimensions[row].height = 22
    data_row(ws4, row, list(item), alt)
    c = ws4.cell(row=row, column=5)
    c.fill = hfill("D5F5E3"); c.font = bold(10,"1E8449"); c.alignment = center()
    alt = not alt
    row += 1

add_dropdown(ws4, "E", 2, row - 1, '"Yes,Skip,Later"')


# ══════════════════════════════════════════════════════════════════
# SHEET 5 – UI/UX NEW DESIGN SPEC
# ══════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("5. UI-UX New Design Spec")
ws5.freeze_panes = "A3"
ws5.sheet_view.showGridLines = False

for letter, width in [("A",24),("B",62),("C",14)]:
    ws5.column_dimensions[letter].width = width

header_row(ws5, 1, ["Component","New Design Spec","Approve"], DARK, height=30)

ux_items = [
    ("/home route  (= Dashboard.jsx)",
     "After login, user lands on Dashboard.jsx served at ROUTES.HOME = '/home'. "
     "home.jsx (the unused file) is deleted. One home screen, no confusion.",
     ""),
    ("Dashboard — KPI Row",
     "5 tiles in one row: Today Orders (blue) | Ready to Deliver (green) | Overdue (red, alert if >0) | "
     "Collected Today Rs. (green) | Pending Payment Rs. (orange). Tap any tile to filter.",
     ""),
    ("Dashboard — Needs Attention",
     "Max 5 items. Each shows: order# + customer + current stage + days stuck + [Move Stage] button. "
     "Red background if overdue. Yellow if 1-2 days late. "
     "Empty state message: All on track!",
     ""),
    ("Dashboard — Stage Pipeline Bar",
     "Horizontal row: Enquiry(n) > Quoted(n) > Approved(n) > Design(n) > Printing(n) > "
     "Post-Print(n) > Ready(n) > Delivered(n). Tap stage to filter Kanban.",
     ""),
    ("Left Sidebar",
     "6 items ONLY: Dashboard | Orders | Customers | Accounts | Reports | Settings (admin). "
     "Flat list, icon + label. No groups, no collapsible sections. Width 200px.",
     ""),
    ("Top Navbar",
     "Left: hamburger + logo. Center: page title. "
     "Right: [+ New Order] orange primary button + bell notification icon + user avatar.",
     ""),
    ("Right Sidebar",
     "DELETED — its 5 quick actions move to top navbar.",
     ""),
    ("Mobile Bottom Nav",
     "4 tabs: Dashboard | Orders | + (add order, prominent) | Accounts. "
     "Sticky bottom. Remove Tasks and Chat tabs.",
     ""),
    ("Add Order Form",
     "6 fields: Customer (autocomplete search), Items+Qty+Rate inline table (max 5 rows), "
     "Due Date (datepicker, default today+3), Priority (Low/Normal/Urgent chip selector), "
     "Advance Paid (toggle yes/no + amount field). "
     "Bottom buttons: [Save as Enquiry] (grey outline) and [Save as Order] (primary orange).",
     ""),
    ("Kanban Board",
     "8 columns (all pipeline stages). Card shows: #N Customer | Item summary | "
     "Due: X days ago/from now | Rs. amount | [Next Stage] button. "
     "Tap card = order detail modal. No drag-drop — stage buttons only.",
     ""),
    ("Order Detail Modal",
     "Header: order#, customer, created date (read-only). "
     "Editable: due date, priority, add note, items qty/rate. "
     "Bottom: stage history timeline. [Mark Paid] button (if stage=delivered). [Print Invoice] button.",
     ""),
    ("Accounts Page",
     "Two tabs: Income | Expenses. "
     "Table: Date | Category | Amount | Order# | Note. "
     "Top summary bar: Income Rs.X — Expenses Rs.Y = Profit Rs.Z. "
     "[+ Add Expense] button. Income entries created automatically when order is paid.",
     ""),
    ("Reports — Orders",
     "Filters: Date Range + Stage dropdown + Customer search. "
     "Table: Order# | Customer | Items | Amount | Stage | Due Date | Assigned To. "
     "Export Excel button top-right.",
     ""),
    ("Reports — Collections",
     "Filters: Date Range + Type (Income/Expense). "
     "Summary row: Total In | Total Out | Net. "
     "Table: Date | Type | Category | Amount | Order# | Note. Export Excel.",
     ""),
    ("Notifications Bell",
     "Dropdown list in top navbar: overdue orders, payment reminders, stage changes. "
     "Badge count. [Mark all read] button.",
     ""),
    ("Colour Theme",
     "Primary: #2C3E50 (dark navy). Accent/CTA: #E67E22 (orange). "
     "Success: #27AE60. Danger: #C0392B. Background: #F8F9FA. Cards: white + subtle shadow.",
     ""),
    ("Typography & Spacing",
     "Font: Inter or system-ui. Headings: 600 weight 16-20px. Body: 400 weight 14px. "
     "Captions: 12px grey. Card padding 16px. Table row height 44px. Mobile tap targets min 44px.",
     ""),
]

row = 2
alt = False
for comp, spec, approve in ux_items:
    ws5.row_dimensions[row].height = 55
    bg = LGREY if alt else WHITE
    for col, val in enumerate([comp, spec, approve], 1):
        c = ws5.cell(row=row, column=col, value=val)
        c.fill   = hfill(bg)
        c.border = thin_border()
        c.alignment = left()
        c.font = bold(10, DARK) if col == 1 else normal()
        if col == 3:
            c.fill = hfill("D5F5E3")
            c.font = bold(10,"1E8449")
            c.alignment = center()
    alt = not alt
    row += 1

add_dropdown(ws5, "C", 2, row - 1, '"Yes,Skip,Later"')


# ══════════════════════════════════════════════════════════════════
# SHEET 6 – EXECUTION ROADMAP
# ══════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("6. Execution Roadmap")
ws6.freeze_panes = "A3"
ws6.sheet_view.showGridLines = False

for letter, width in [("A",6),("B",30),("C",22),("D",48),("E",16),("F",14),("G",16)]:
    ws6.column_dimensions[letter].width = width

header_row(ws6, 1,
    ["Phase","Phase Name","Files Changed","Tasks","Est. Effort","Approve","Status"],
    DARK, height=30)

roadmap = [
    ("Phase 1 — Delete",),
    ("1", "Delete dead code",
     "~26 FE pages, ~12 BE routes, ~8 DB models",
     "Delete home.jsx (unrouted). Remove Baileys, Gmail, FlowBuilder, BankReconciliation, "
     "TrialBalance, CallLogs, DiaryUpload, legacy WA pages/components, 14 reports, "
     "SopPage, WorkflowTemplates, PurchaseOrder. Update App.jsx routes.",
     "1-2 days", "", "Pending"),

    ("Phase 2 — Navigation",),
    ("2", "Rebuild Sidebar + Navbar + Footer",
     "Sidebar.jsx, TopNavbar.jsx, Footer.jsx, delete sidebarMenu.jsx",
     "Flat 6-item sidebar, top quick-action bar with [+ New Order], 4-tab mobile footer.",
     "0.5 day", "", "Pending"),

    ("Phase 3 — Dashboard  (= /home)",),
    ("3", "Rebuild Dashboard.jsx  (the /home route)",
     "Dashboard.jsx, Components/dashboard/ folder",
     "5 KPI tiles + Needs Attention list + Stage pipeline bar. "
     "Remove resizable panels, customization drawer, attendance section, user tasks section.",
     "1 day", "", "Pending"),

    ("Phase 4 — Order Forms",),
    ("4", "Rebuild Add Order + Order Update",
     "addOrder1.jsx, OrderUpdate.jsx",
     "6-field add order form. Simplified edit form (read-only + editable fields only).",
     "1 day", "", "Pending"),

    ("Phase 5 — Kanban",),
    ("5", "Simplify Kanban board",
     "OrderKanban.jsx, OrderBoard.jsx, OrderCard.jsx",
     "Simpler cards, [Next Stage] button, 8 columns, remove drag-drop.",
     "0.5 day", "", "Pending"),

    ("Phase 6 — Reports",),
    ("6", "Rebuild Reports  (17 -> 3)",
     "Reports/ folder",
     "Orders Report, Collections Report (new file), Customer Report (minor cleanup).",
     "1 day", "", "Pending"),

    ("Phase 7 — Accounts",),
    ("7", "Simplify Accounts",
     "Pages/DayBook.jsx, accounts backend routes",
     "Replace double-entry DayBook with simple Income/Expense log. Two tabs.",
     "0.5 day", "", "Pending"),

    ("Phase 8 — Backend Automation",),
    ("8", "Add automation",
     "orderLifecycleController.js, messageScheduler.js, whatsappTemplates.js, orderService.js",
     "Auto WA on order creation + stage changes. Payment reminders cron (10am). "
     "Daily owner summary cron (7pm). Auto income entry on payment.",
     "1 day", "", "Pending"),

    ("Phase 9 — DB Cleanup",),
    ("9", "Clean DB models",
     "repositories/ folder + migration script",
     "Remove unused collections (Enquiry, Flow, Baileys, Gmail, BankStatement, PurchaseOrder, CallLogs, DiaryDraft). "
     "Remove Order.Status[] array. Add pagination to /api/orders/all-data.",
     "0.5 day", "", "Pending"),

    ("Phase 10 — QA",),
    ("10", "End-to-end testing",
     "All changed files",
     "Golden path: Create order (target <2 min) > advance stages > verify WA sent > "
     "mark paid > check collections report. Test mobile layout. Check role-based access.",
     "0.5 day", "", "Pending"),
]

row = 2
alt = False
for item in roadmap:
    if len(item) == 1:
        section_header(ws6, row, f"  ▶  {item[0]}", DARK, 7)
        row += 1
        continue
    ws6.row_dimensions[row].height = 55
    bg = LGREY if alt else WHITE
    for col, val in enumerate(list(item), 1):
        c = ws6.cell(row=row, column=col, value=val)
        c.fill   = hfill(bg)
        c.font   = normal()
        c.border = thin_border()
        c.alignment = left()
    c6 = ws6.cell(row=row, column=6)
    c6.fill = hfill("D5F5E3"); c6.font = bold(10,"1E8449"); c6.alignment = center()
    c7 = ws6.cell(row=row, column=7)
    c7.fill = hfill("EBF5FB"); c7.font = normal(10,"2471A3"); c7.alignment = center()
    alt = not alt
    row += 1

add_dropdown(ws6, "F", 2, row - 1, '"Yes,Skip,Later"')
add_dropdown(ws6, "G", 2, row - 1, '"Pending,In Progress,Done,On Hold"')

# ── save ─────────────────────────────────────────────────────────
out = "/home/user/MIS-Both-/MIS_Reconstruction_Plan.xlsx"
wb.save(out)
print(f"Saved: {out}")
